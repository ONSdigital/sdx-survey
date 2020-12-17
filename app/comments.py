import os
import sys
from string import ascii_lowercase
from google.cloud import datastore

from app.encryption import encrypt_survey

parent_dir_path = os.path.dirname(os.path.dirname(os.path.realpath(__file__)))
sys.path.append(parent_dir_path)

from openpyxl import Workbook

datastore_client = datastore.Client()


# def create_comments_excel_file(survey_id, period, submissions):
#     """Extract comments from submissions and write them to an excel file"""
#     print("Generating Excel file")
#     workbook = Workbook()
#     row = 2
#     surveys_with_comments_count = 0
#     ws = workbook.active
#
#     for submission in submissions:
#         comment = get_comment_text(submission)
#
#         boxes_selected = ""
#         for key in ('146' + letter for letter in ascii_lowercase[0:]):
#             if key in submission.data['data'].keys():
#                 boxes_selected = boxes_selected + key + ' '
#
#         if not comment:
#             continue
#         row += 1
#         surveys_with_comments_count += 1
#         ws.cell(row, 1, submission.data['metadata']['ru_ref'])
#         ws.cell(row, 2, submission.data['collection']['period'])
#         if survey_id == '134':
#             if '300w' in submission.data['data']:
#                 ws.cell(row, 5, submission.data['data']['300w'])
#             if '300f' in submission.data['data']:
#                 ws.cell(row, 6, submission.data['data']['300f'])
#             if '300m' in submission.data['data']:
#                 ws.cell(row, 7, submission.data['data']['300m'])
#             if '300w4' in submission.data['data']:
#                 ws.cell(row, 8, submission.data['data']['300w4'])
#             if '300w5' in submission.data['data']:
#                 ws.cell(row, 9, submission.data['data']['300w5'])
#             checkboxes = ['91w', '92w1', '92w2', '94w1', '94w2', '95w', '96w', '97w',
#                           '91f', '92f1', '92f2', '94f1', '94f2', '95f', '96f', '97f',
#                           '191m', '192m1', '192m2', '194m1', '194m2', '195m', '196m', '197m',
#                           '191w4', '192w41', '192w42', '194w41', '194w42', '195w4', '196w4', '197w4',
#                           '191w5', '192w51', '192w52', '194w51', '194w52', '195w5', '196w5', '197w5']
#             for checkbox in checkboxes:
#                 if checkbox in submission.data['data']:
#                     boxes_selected = boxes_selected + f"{checkbox}, "
#         ws.cell(row, 3, boxes_selected)
#         ws.cell(row, 4, comment)
#
#     ws.cell(1, 1, f"Survey ID: {survey_id}")
#     ws.cell(1, 2, f"Comments found: {surveys_with_comments_count}")
#     if survey_id == '134':
#         ws.cell(1, 5, "Weekly comment")
#         ws.cell(1, 6, "Fortnightly comment")
#         ws.cell(1, 7, "Calendar Monthly comment")
#         ws.cell(1, 8, "4 Weekly Pay comment")
#         ws.cell(1, 9, "5 Weekly Pay comment")
#     print(f"{surveys_with_comments_count} out of {len(submissions)} submissions had comments")
#
#     parent_dir_path = os.path.dirname(os.path.dirname(os.path.realpath(__file__)))
#     filename = os.path.join(parent_dir_path, f"{survey_id}_{period}.xlsx")
#     workbook.save(filename)
#     workbook.close()
#     print(f"Excel file {filename} generated")


def store_comments(survey_dict: dict) -> str:
    transaction_id = survey_dict["tx_id"]
    period = survey_dict["collection"]["period"]
    survey_id = survey_dict["survey_id"]
    data = {"ru_ref": survey_dict["metadata"]["ru_ref"],
            "boxes_selected": get_boxes_selected(survey_id),
            "comments": get_comments_list()}
    encrypted_data = encrypt_survey(data)

    comment = Comment(transaction_id=transaction_id,
                      period=period,
                      survey_id=survey_id,
                      encrypted_data=encrypted_data)

    commit_to_datastore(comment)


def get_comments_list(submission: dict):
    """Returns the respondent typed text from a submission.  The qcode for this text will be different depending
    on the survey
    """
    comments_list = []
    if submission['survey_id'] == '134':
        if '300w' in submission['data']:
            comments_list.append(get_comment(submission, '300w'))
        if '300f' in submission['data']:
            comments_list.append(get_comment(submission, '300f'))
        if '300m' in submission['data']:
            comments_list.append(get_comment(submission, '300m'))
        if '300w4' in submission['data']:
            comments_list.append(get_comment(submission, '300w4'))
        if '300w5' in submission['data']:
            comments_list.append(get_comment(submission, '300w5'))

    else:
        if submission['survey_id'] == '187':
            comments_list.append(get_comment(submission, '500'))
        elif submission['survey_id'] == '134':
            comments_list.append(get_comment(submission, '300'))
        else:
            comments_list.append(get_comment(submission, '146'))

    return comments_list


def get_comment(submission, qcode):
    return submission['data'].get(qcode)


def get_boxes_selected(submission):
    boxes_selected = ''
    if submission['survey_id'] == '134':
        checkboxes = ['91w', '92w1', '92w2', '94w1', '94w2', '95w', '96w', '97w',
                      '91f', '92f1', '92f2', '94f1', '94f2', '95f', '96f', '97f',
                      '191m', '192m1', '192m2', '194m1', '194m2', '195m', '196m', '197m',
                      '191w4', '192w41', '192w42', '194w41', '194w42', '195w4', '196w4', '197w4',
                      '191w5', '192w51', '192w52', '194w51', '194w52', '195w5', '196w5', '197w5']
        for checkbox in checkboxes:
            if checkbox in submission.data['data']:
                boxes_selected = boxes_selected + f"{checkbox}, "

    else:
        for key in ('146' + letter for letter in ascii_lowercase[0:]):
            if key in submission.keys():
                boxes_selected = boxes_selected + key + ' '

    return boxes_selected


class Comment:
    def __init__(self, transaction_id, survey_id, period, encrypted_data):
        self.transaction_id = transaction_id
        self.survey_id = survey_id
        self.period = period
        self.encrypted_data = encrypted_data
        self.zip_name = None


def commit_to_datastore(comment):
    try:
        entity_key = datastore_client.key('Comment', comment.transaction_id)
        entity = datastore.Entity(key=entity_key)
        entity.update(
            {
                "survey_id": comment.survey_id,
                "period": comment.period,
                "zip_name": comment.zip_name,
                "encrypted_data": comment.encrypted_data
            }
        )
        return datastore.Client().put(entity)
    except ValueError as error:
        print(error)
